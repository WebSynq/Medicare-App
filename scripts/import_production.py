#!/usr/bin/env python3
"""
import_production.py
====================
Import the Plecto "GHW Production Tracker" into the production_records
MongoDB collection. Accepts either an .xlsx file directly (the native
Plecto / shared-drive shape — e.g. ROOKIE_SEASON_DOC.xlsx) or a .csv
export. The file extension determines the reader; the row normalization
and upsert logic are identical for both.

The script is intentionally permissive about column naming because Plecto
exports drift over time and we don't want a column rename to silently drop
rows. Required columns are looked up by a list of candidate names; missing
required fields skip the row with a logged reason.

Usage:
    python scripts/import_production.py /path/to/ROOKIE_SEASON_DOC.xlsx
    python scripts/import_production.py /path/to/export.csv

Env vars:
    MONGO_URL  (required) — Atlas connection string
    DB_NAME    (required) — target database

Behaviour:
- For .xlsx files: reads the first (active) worksheet, row 1 is the header.
  Cells already typed by Excel (numbers, dates) are coerced to string before
  passing through the same normalizers used for CSV — keeps one source of
  truth for parsing.
- Premium is normalized: "$1,801.00" / "1801" / "1,801" → 1801.0 float
- Revenue (commission) is normalized the same way → revenue_expected
- Dates are parsed permissively (MM/DD/YYYY, YYYY-MM-DD, ISO, etc.).
  Excel-typed datetime cells are converted via .date().isoformat().
- Agent identity is resolved via the email-prefix map below; rows whose
  email prefix is unknown are imported with agent_name=None and logged.
- Idempotency: each row's natural key (carrier|policy_number|effective_date|
  agent_email) is hashed; insert is upserted on that hash so re-running
  the script is safe and won't double-count.
- revenue_received and audit_status_at_received are not populated here —
  they are set later when AgencyBloc sync confirms receipt.
"""
from __future__ import annotations

import argparse
import asyncio
import csv
import hashlib
import logging
import os
import re
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Iterator, Optional

# Make backend/ importable so we can reuse the motor client config.
ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "backend"))

from motor.motor_asyncio import AsyncIOMotorClient  # noqa: E402


logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)s %(message)s",
)
logger = logging.getLogger("import_production")


# ── Agent identity map ──────────────────────────────────────────────────────
# Email prefix (everything before @) → canonical agent_name. Lowercase keys.
# Source: confirmed by the GHW team. New agents must be added here AND
# invited via the /auth/invite admin flow so their user row has agent_name.
EMAIL_PREFIX_TO_AGENT = {
    "wlunt": "Wes Lunt",
    "tgoveia": "Travis Goveia",
    "jhockaday": "Jack Hockaday",
    "nstevens": "Nic Stevens",
    "asmith": "Ron Smith",
    "chunt": "Chris Hunt",
    "fchildress": "Feleshia Childress",
    "cgruening": "Leadership (Chase Gruening)",
    "dcurtis": "Devin Curtis",
    "bryce": "Bryce Pritchard",
    "conor": "Conor McCormick",
    "mhughes": "Meagan Hughes",
    "tim": "Tim Dazey",
    "summer": "Summer Bell",
    "kyle": "Kyle Welch",
    "levi": "Levi Plaster",
    "austin": "Austin Compton",
    "matt": "Matt Monacelli",
    "ethan": "Ethan Hinds",
}


# ── Column resolution ──────────────────────────────────────────────────────
# Multiple candidate header names per logical field. First match wins.
# All comparisons are case-insensitive and ignore surrounding whitespace.
COLUMN_CANDIDATES = {
    "agent_email":      ["Agent Email", "Email", "Agent E-mail", "agent_email"],
    "agent_name_raw":   ["Agent Name", "Agent", "Producer", "agent_name"],
    "policy_number":    ["Policy Number", "Policy #", "Policy", "PolicyNumber"],
    "client_name":      ["Client Name", "Client", "Insured Name", "Member Name"],
    "carrier":          ["Carrier", "Company", "Insurance Company"],
    "product":          ["Product", "Plan", "Plan Type", "Product Type"],
    "state":            ["State", "Member State", "Client State"],
    "effective_date":   ["Effective Date", "Eff Date", "Effective", "Policy Effective"],
    "submitted_date":   ["Submitted Date", "Submit Date", "Submitted"],
    "premium":          ["Premium", "Monthly Premium", "Premium Amount"],
    "revenue":          ["Revenue", "Commission", "Expected Commission",
                         "Revenue Expected", "Commission Amount"],
    "ab_flag":          ["Enter in AB", "AgencyBloc", "AB", "Synced to AB"],
}


def _normalize_header(h: str) -> str:
    return h.strip().lower()


def _resolve_columns(header_row: list[str]) -> dict[str, Optional[str]]:
    """Map logical name → actual CSV column name (preserving original case)
    so we can re-read each row by the original key.
    """
    lookup = {_normalize_header(h): h for h in header_row}
    out: dict[str, Optional[str]] = {}
    for logical, candidates in COLUMN_CANDIDATES.items():
        found = None
        for cand in candidates:
            if _normalize_header(cand) in lookup:
                found = lookup[_normalize_header(cand)]
                break
        out[logical] = found
    return out


# ── Field normalisers ──────────────────────────────────────────────────────
_MONEY_RE = re.compile(r"[^0-9.\-]")


def _normalize_money(raw: Optional[str]) -> Optional[float]:
    """Strip $, commas, whitespace. Empty / None → None.

    "$1,801.00" → 1801.0
    "1801"      → 1801.0
    ""          → None
    """
    if raw is None:
        return None
    s = str(raw).strip()
    if not s:
        return None
    s = _MONEY_RE.sub("", s)
    if not s or s in ("-", ".", "-."):
        return None
    try:
        return float(s)
    except ValueError:
        return None


# Date format candidates tried in order. Python 3.11.9 compatible (no fromisoformat
# tricks that only land in 3.12).
_DATE_FORMATS = (
    "%Y-%m-%d",
    "%m/%d/%Y",
    "%m/%d/%y",
    "%m-%d-%Y",
    "%d/%m/%Y",
    "%Y/%m/%d",
    "%b %d, %Y",
    "%B %d, %Y",
)


def _parse_date(raw: Optional[str]) -> Optional[str]:
    """Return ISO-8601 date (YYYY-MM-DD) or None.

    We persist as ISO string rather than BSON date so the field is
    JSON-roundtrippable and matches the lead/SOA convention in this app.
    """
    if raw is None:
        return None
    s = str(raw).strip()
    if not s:
        return None
    # ISO with time
    try:
        return datetime.fromisoformat(s.replace("Z", "+00:00")).date().isoformat()
    except ValueError:
        pass
    for fmt in _DATE_FORMATS:
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            continue
    return None


def _normalize_bool_flag(raw: Optional[str]) -> Optional[bool]:
    if raw is None:
        return None
    s = str(raw).strip().lower()
    if not s:
        return None
    if s in ("yes", "y", "true", "1", "x", "✓", "✔"):
        return True
    if s in ("no", "n", "false", "0"):
        return False
    return None


def _agent_name_from_email(email: Optional[str]) -> Optional[str]:
    if not email:
        return None
    s = email.strip().lower()
    if "@" not in s:
        return None
    prefix = s.split("@", 1)[0]
    return EMAIL_PREFIX_TO_AGENT.get(prefix)


def _natural_key_hash(carrier: str, policy_number: str,
                       effective_date: Optional[str],
                       agent_email: Optional[str]) -> str:
    """Stable hash of the row's natural identifiers. Used as the upsert key
    so re-imports don't duplicate. Trailing whitespace and case are
    normalised so spreadsheet edits don't break idempotency."""
    parts = [
        (carrier or "").strip().lower(),
        (policy_number or "").strip().lower(),
        (effective_date or "").strip().lower(),
        (agent_email or "").strip().lower(),
    ]
    return hashlib.sha256("|".join(parts).encode("utf-8")).hexdigest()


# ── Row → document ──────────────────────────────────────────────────────────
def _row_to_doc(row: dict, cols: dict[str, Optional[str]]) -> Optional[dict]:
    """Translate one CSV row into a production_record document.

    Returns None if the row is missing a natural-key field — the caller
    counts these as skipped rows.
    """
    def _get(logical: str) -> Optional[str]:
        actual = cols.get(logical)
        if not actual:
            return None
        v = row.get(actual)
        return v.strip() if isinstance(v, str) else v

    agent_email = _get("agent_email")
    policy_number = _get("policy_number")
    carrier = _get("carrier")
    effective_date = _parse_date(_get("effective_date"))

    if not policy_number or not carrier:
        return None  # cannot dedupe without both — skip and log

    premium_monthly = _normalize_money(_get("premium"))
    premium_annual = round(premium_monthly * 12, 2) if premium_monthly is not None else None
    revenue_expected = _normalize_money(_get("revenue"))

    # Prefer the agent_name we resolve from email; fall back to whatever the
    # CSV's Agent Name column said (rare — for legacy rows where email is
    # missing or a contractor without a portal account).
    agent_name = _agent_name_from_email(agent_email) or _get("agent_name_raw")

    now_iso = datetime.now(timezone.utc).isoformat()
    return {
        "natural_key": _natural_key_hash(carrier, policy_number, effective_date, agent_email),
        "agent_email": (agent_email or "").lower() or None,
        "agent_name": agent_name,
        "policy_number": policy_number,
        "client_name": _get("client_name"),
        "carrier": carrier,
        "product": _get("product"),
        "state": (_get("state") or "").upper() or None,
        "effective_date": effective_date,
        "submitted_date": _parse_date(_get("submitted_date")),
        "premium_monthly": premium_monthly,
        "premium_annual": premium_annual,
        "revenue_expected": revenue_expected,
        "revenue_received": None,
        "audit_status": "pending",
        "audit_notes": None,
        "ab_synced": _normalize_bool_flag(_get("ab_flag")),
        "imported_at": now_iso,
        "updated_at": now_iso,
    }


# ── File readers (CSV / XLSX) ───────────────────────────────────────────────
def _cell_to_str(value) -> str:
    """Coerce one openpyxl cell value to the string shape the CSV path produces.

    openpyxl returns native Python types for typed cells (datetime, int, float,
    bool). Stringifying here means the downstream normalizers (_normalize_money,
    _parse_date, _normalize_bool_flag) only need to handle one input shape.
    Empty cells (None) become "" so column resolution still finds the key.
    """
    if value is None:
        return ""
    # Preserve ISO-date precision instead of going through datetime.__str__,
    # which prepends 'T00:00:00' for plain dates.
    from datetime import date, datetime as _dt
    if isinstance(value, _dt):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        # "1801.0" → "1801" so the money regex doesn't trip on a trailing zero
        return str(int(value))
    return str(value).strip()


def _iter_csv(path: Path) -> tuple[list[str], Iterator[dict]]:
    """Open a CSV and return (header_list, dict-row iterator).

    The caller is responsible for keeping the iterator scoped to a `with`
    block; we expose the iterator directly so the upsert loop streams the
    file rather than materializing it.
    """
    fh = path.open("r", encoding="utf-8-sig", newline="")
    reader = csv.DictReader(fh)
    if not reader.fieldnames:
        fh.close()
        raise SystemExit(f"{path}: no header row.")
    headers = list(reader.fieldnames)

    def _gen():
        try:
            for row in reader:
                yield row
        finally:
            fh.close()

    return headers, _gen()


def _iter_xlsx(path: Path) -> tuple[list[str], Iterator[dict]]:
    """Open an XLSX (first sheet) and yield dict rows shaped like csv.DictReader.

    read_only=True streams the workbook (constant memory for large files);
    data_only=True returns evaluated values for cells that contain formulas
    so Excel-computed totals come through as numbers, not '=SUM(...)' strings.
    """
    try:
        from openpyxl import load_workbook  # local import — script can still run on a CSV without openpyxl installed
    except ImportError as e:
        raise SystemExit(
            "openpyxl is required to read .xlsx files. "
            "Install with: pip install openpyxl"
        ) from e

    wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    ws = wb.active  # first / active sheet — matches Plecto's single-sheet export
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        wb.close()
        raise SystemExit(f"{path}: workbook has no rows.")

    # Drop trailing empty header columns (Excel often has phantom None cells
    # past the last filled column) and coerce headers to plain strings.
    headers = [_cell_to_str(h) for h in header_row]
    while headers and headers[-1] == "":
        headers.pop()
    if not headers:
        wb.close()
        raise SystemExit(f"{path}: first row has no headers.")

    def _gen():
        try:
            for row_values in rows_iter:
                # Skip completely empty rows — Plecto sometimes leaves blank
                # rows between sections in shared trackers.
                if not any(v is not None and str(v).strip() != ""
                            for v in row_values[:len(headers)]):
                    continue
                yield {
                    headers[i]: _cell_to_str(row_values[i])
                    for i in range(len(headers))
                    if i < len(row_values)
                }
        finally:
            wb.close()

    return headers, _gen()


def _open_rows(path: Path) -> tuple[list[str], Iterator[dict]]:
    """Dispatch to the right reader based on file extension."""
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return _iter_csv(path)
    if suffix in (".xlsx", ".xlsm"):
        return _iter_xlsx(path)
    raise SystemExit(
        f"Unsupported file extension: {suffix!r}. Expected .csv or .xlsx."
    )


# ── Main ────────────────────────────────────────────────────────────────────
async def _import(source_path: Path) -> dict:
    mongo_url = os.environ.get("MONGO_URL")
    db_name = os.environ.get("DB_NAME")
    if not mongo_url or not db_name:
        raise SystemExit("MONGO_URL and DB_NAME env vars are required.")

    client = AsyncIOMotorClient(mongo_url)
    db = client[db_name]
    coll = db.production_records

    # natural_key is the upsert key; ensure the unique index exists before
    # we touch any rows. Idempotent — Mongo no-ops if the index is already
    # there.
    await coll.create_index("natural_key", unique=True)
    await coll.create_index("agent_email")
    await coll.create_index("agent_name")
    await coll.create_index("effective_date")
    await coll.create_index("audit_status")

    inserted = 0
    updated = 0
    skipped_no_key = 0
    skipped_no_agent = 0
    total = 0

    headers, row_iter = _open_rows(source_path)
    cols = _resolve_columns(headers)
    missing = [k for k in ("agent_email", "carrier", "policy_number", "revenue")
               if cols.get(k) is None]
    if missing:
        logger.warning(
            "%s is missing recommended columns: %s. Available headers: %s",
            source_path.name, missing, headers,
        )

    for row in row_iter:
        total += 1
        doc = _row_to_doc(row, cols)
        if doc is None:
            skipped_no_key += 1
            continue
        if doc["agent_name"] is None:
            skipped_no_agent += 1
            logger.warning("Row %d: unknown agent email %r — keeping with agent_name=null",
                            total, doc.get("agent_email"))

        result = await coll.update_one(
            {"natural_key": doc["natural_key"]},
            {
                "$set": {
                    # Refreshable fields — always overwrite from source
                    "agent_email": doc["agent_email"],
                    "agent_name": doc["agent_name"],
                    "client_name": doc["client_name"],
                    "product": doc["product"],
                    "state": doc["state"],
                    "submitted_date": doc["submitted_date"],
                    "premium_monthly": doc["premium_monthly"],
                    "premium_annual": doc["premium_annual"],
                    "revenue_expected": doc["revenue_expected"],
                    "ab_synced": doc["ab_synced"],
                    "updated_at": doc["updated_at"],
                },
                # Set-once fields — preserve across re-imports
                "$setOnInsert": {
                    "natural_key": doc["natural_key"],
                    "policy_number": doc["policy_number"],
                    "carrier": doc["carrier"],
                    "effective_date": doc["effective_date"],
                    "revenue_received": None,
                    "audit_status": "pending",
                    "audit_notes": None,
                    "imported_at": doc["imported_at"],
                },
            },
            upsert=True,
        )
        if result.upserted_id is not None:
            inserted += 1
        elif result.modified_count:
            updated += 1

    client.close()
    return {
        "total_rows": total,
        "inserted": inserted,
        "updated": updated,
        "skipped_no_key": skipped_no_key,
        "skipped_no_agent": skipped_no_agent,
    }


def main() -> None:
    parser = argparse.ArgumentParser(
        description=(
            "Import the Plecto GHW Production Tracker into production_records. "
            "Accepts .xlsx (e.g. ROOKIE_SEASON_DOC.xlsx) or .csv. "
            "Idempotent — safe to re-run."
        ),
    )
    parser.add_argument(
        "source_path",
        type=Path,
        help="Path to the Plecto tracker file (.xlsx or .csv).",
    )
    args = parser.parse_args()
    if not args.source_path.exists():
        raise SystemExit(f"File not found: {args.source_path}")

    stats = asyncio.run(_import(args.source_path))
    logger.info(
        "Import done: total=%d inserted=%d updated=%d skipped_no_key=%d skipped_no_agent=%d",
        stats["total_rows"], stats["inserted"], stats["updated"],
        stats["skipped_no_key"], stats["skipped_no_agent"],
    )


if __name__ == "__main__":
    main()
